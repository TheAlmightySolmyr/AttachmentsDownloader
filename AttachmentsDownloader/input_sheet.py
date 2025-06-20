import openpyxl

class GetAlongWithSheet:
    def __init__(self, sheet_path):
        self.sheet = openpyxl.load_workbook(sheet_path)

    def activator(self):
        self.sheet.active
    
    def get_sheets_names(self):
        return self.sheet.sheetnames
    
    def switch_list(self, page):
        return self.sheet[page]
    
    def cell_value(self, row, column):
        return self.sheet.cell(row,column).value